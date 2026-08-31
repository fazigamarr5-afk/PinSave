import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors
HEADER_FILL = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
PILLAR_FILL = PatternFill(start_color="FDF2F8", end_color="FDF2F8", fill_type="solid")
SECTION_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
HIGH_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
MED_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
LOW_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
BOLD = Font(name="Arial", bold=True, size=11)
NORMAL = Font(name="Arial", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = THIN_BORDER

# ============================================================
# SHEET 1: Content Strategy Overview
# ============================================================
ws1 = wb.active
ws1.title = "Strategy Overview"
ws1.sheet_properties.tabColor = "E11D48"

# Title
ws1.merge_cells("A1:F1")
ws1.cell(row=1, column=1, value="SavePin Content Strategy — npftas.xyz").font = Font(name="Arial", bold=True, size=16, color="E11D48")
ws1.cell(row=2, column=1, value="3-Month Editorial Calendar | Target: Pinterest Downloader Keywords").font = Font(name="Arial", size=11, color="64748B")
ws1.cell(row=3, column=1, value="Last Updated: August 2026").font = Font(name="Arial", size=10, color="94A3B8")

# Content Pillars
ws1.cell(row=5, column=1, value="CONTENT PILLARS").font = BOLD
headers = ["Pillar", "Description", "Target Audience", "Content Types", "Connection to Product"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=6, column=i, value=h)
style_header(ws1, 6, 5)

pillars = [
    ["Pinterest Downloading", "How-to guides, tutorials, and tool comparisons for downloading Pinterest content", "Users searching for Pinterest downloaders", "How-to, Comparison, Tutorial", "Core product — drives direct tool usage"],
    ["Pinterest Tips & Tricks", "Best practices for using Pinterest effectively, finding content, organizing pins", "Active Pinterest users", "Guide, Listicle, Tip Roundup", "Builds authority and attracts Pinterest power users"],
    ["Content Creation Tools", "Reviews and comparisons of tools for creating, editing, and repurposing Pinterest content", "Content creators and marketers", "Comparison, Review, Guide", "Positions SavePin as part of the creator toolkit"],
    ["Pinterest for Business", "Using Pinterest for marketing, traffic, and brand growth", "Small business owners, marketers", "Strategy, Case Study, Data", "Attracts business users who need reliable download tools"],
    ["Legal & Ethics", "Copyright, fair use, and responsible content sharing on Pinterest", "Concerned users, creators", "FAQ, Guide, Opinion", "Builds trust and differentiates from shady competitors"],
]

for i, row in enumerate(pillars, 7):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
    style_data(ws1, i, 5)
    ws1.cell(row=i, column=1).fill = PILLAR_FILL
    ws1.cell(row=i, column=1).font = BOLD

# Calendar Split
ws1.cell(row=13, column=1, value="CALENDAR SPLIT (60/30/10)").font = BOLD
split_headers = ["Type", "Percentage", "Monthly Posts", "Description"]
for i, h in enumerate(split_headers, 1):
    ws1.cell(row=14, column=i, value=h)
style_header(ws1, 14, 4)

split_data = [
    ["Searchable", "60%", "3", "How-to guides, comparisons, tutorials — captures existing search demand"],
    ["Shareable", "30%", "2", "Original data, thought leadership, stats roundups — creates demand"],
    ["Experimental", "10%", "0-1", "New formats, interactive tools, video scripts — cheap insurance"],
]

for i, row in enumerate(split_data, 15):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
    style_data(ws1, i, 4)

# Column widths
for col, width in enumerate([25, 45, 35, 30, 40], 1):
    ws1.column_dimensions[get_column_letter(col)].width = width

# ============================================================
# SHEET 2: Keyword Research & Priorities
# ============================================================
ws2 = wb.create_sheet("Keywords & Priorities")
ws2.sheet_properties.tabColor = "22C55E"

headers2 = ["Keyword", "Monthly Volume", "Difficulty", "Buyer Stage", "Content Type", "Priority", "Status"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, 7)

keywords = [
    # High priority — Awareness
    ["pinterest video downloader", 49500, "Medium", "Awareness", "Tool Page", "P0 — Do First", "Live"],
    ["download pinterest videos", 33100, "Medium", "Awareness", "How-To Guide", "P0 — Do First", "Draft"],
    ["pinterest downloader", 27100, "High", "Awareness", "Tool Page", "P0 — Do First", "Live"],
    ["how to download pinterest videos", 22200, "Low", "Awareness", "Tutorial", "P0 — Do First", "Existing"],
    ["pinterest video downloader free", 18100, "Low", "Awareness", "Comparison", "P0 — Do First", "Planned"],
    ["pinterest image downloader", 14800, "Medium", "Awareness", "Tool Page", "P1 — Next", "Live"],
    ["pinterest gif downloader", 9900, "Low", "Awareness", "Tool Page", "P1 — Next", "Live"],
    ["save pinterest videos", 8100, "Low", "Awareness", "How-To", "P1 — Next", "Planned"],

    # Consideration
    ["best pinterest downloader", 6600, "Medium", "Consideration", "Comparison", "P0 — Do First", "Draft"],
    ["pinterest downloader app", 5400, "Medium", "Consideration", "Comparison", "P1 — Next", "Planned"],
    ["pinterest video downloader no watermark", 4400, "Low", "Consideration", "Comparison", "P1 — Next", "Planned"],
    ["pinterest downloader online", 3600, "Low", "Consideration", "Tool Page", "P1 — Next", "Live"],
    ["pinterest to mp4", 2900, "Low", "Consideration", "How-To", "P2 — Later", "Planned"],
    ["pinterest video downloader apk", 2400, "Medium", "Consideration", "Comparison", "P2 — Later", "Planned"],

    # Implementation
    ["how to save pinterest videos to camera roll", 1900, "Low", "Implementation", "Tutorial", "P1 — Next", "Existing"],
    ["how to download pinterest videos on iphone", 1600, "Low", "Implementation", "Tutorial", "P1 — Next", "Planned"],
    ["how to download pinterest videos on android", 1300, "Low", "Implementation", "Tutorial", "P1 — Next", "Planned"],
    ["pinterest video download chrome extension", 1100, "Medium", "Implementation", "Review", "P2 — Later", "Planned"],

    # Long-tail / Shareable
    ["is it legal to download pinterest videos", 880, "Low", "Awareness", "FAQ/Guide", "P1 — Next", "Existing"],
    ["pinterest download limit", 720, "Low", "Implementation", "FAQ", "P2 — Later", "Planned"],
    ["pinterest video quality explained", 590, "Low", "Awareness", "Guide", "P2 — Later", "Planned"],
    ["pinterest vs instagram reels download", 480, "Low", "Consideration", "Comparison", "P2 — Later", "Planned"],
    ["pinterest video downloader chrome", 1200, "Medium", "Implementation", "Review", "P2 — Later", "Planned"],
    ["pinterest downloader extension", 880, "Medium", "Implementation", "Review", "P2 — Later", "Planned"],
]

for i, row in enumerate(keywords, 2):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val)
    style_data(ws2, i, 7)
    # Color by priority
    priority = row[5]
    if "P0" in priority:
        ws2.cell(row=i, column=6).fill = HIGH_FILL
    elif "P1" in priority:
        ws2.cell(row=i, column=6).fill = MED_FILL
    else:
        ws2.cell(row=i, column=6).fill = LOW_FILL

for col, width in enumerate([38, 15, 12, 16, 18, 16, 12], 1):
    ws2.column_dimensions[get_column_letter(col)].width = width

# ============================================================
# SHEET 3: Editorial Calendar (3 Months)
# ============================================================
ws3 = wb.create_sheet("Editorial Calendar")
ws3.sheet_properties.tabColor = "3B82F6"

headers3 = ["Week", "Publish Date", "Title", "Pillar", "Type", "Target Keyword", "Buyer Stage", "Word Count", "Status"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, 9)

calendar = [
    # Month 1 — Foundation
    ["W1", "Sep 8", "SavePin vs Other Pinterest Downloaders: Complete 2026 Comparison", "Pinterest Downloading", "Comparison (Searchable)", "best pinterest downloader", "Consideration", "2500", "Draft Ready"],
    ["W2", "Sep 15", "How to Download Pinterest Videos on iPhone (2026 Guide)", "Pinterest Downloading", "Tutorial (Searchable)", "how to download pinterest videos on iphone", "Implementation", "1500", "Planned"],
    ["W3", "Sep 22", "Pinterest Video Quality Guide: 720p vs 1080p vs 4K", "Pinterest Tips", "Guide (Searchable)", "pinterest video quality explained", "Awareness", "1800", "Planned"],
    ["W4", "Sep 29", "Pinterest Download Statistics 2026: Usage, Trends & Data", "Pinterest Tips", "Data (Shareable)", "pinterest download statistics", "Awareness", "1200", "Planned"],

    # Month 2 — Expansion
    ["W5", "Oct 6", "How to Download Pinterest Videos on Android (Step-by-Step)", "Pinterest Downloading", "Tutorial (Searchable)", "how to download pinterest videos on android", "Implementation", "1500", "Planned"],
    ["W6", "Oct 13", "10 Best Pinterest Downloaders Tested & Ranked (Free Tools)", "Content Tools", "Comparison (Searchable)", "best pinterest downloader free", "Consideration", "2200", "Planned"],
    ["W7", "Oct 20", "Is It Legal to Download Pinterest Videos? Complete Guide", "Legal & Ethics", "FAQ/Guide (Both)", "is it legal to download pinterest videos", "Awareness", "1800", "Planned"],
    ["W8", "Oct 27", "Pinterest Marketing: How to Use Pinterest for Business in 2026", "Pinterest for Business", "Guide (Shareable)", "pinterest marketing 2026", "Awareness", "2500", "Planned"],

    # Month 3 — Authority
    ["W9", "Nov 3", "Pinterest Downloader Chrome Extensions: Do You Need One?", "Content Tools", "Review (Searchable)", "pinterest video downloader chrome", "Implementation", "1500", "Planned"],
    ["W10", "Nov 10", "How to Save Pinterest Videos to Camera Roll (Any Device)", "Pinterest Downloading", "Tutorial (Searchable)", "save pinterest videos to camera roll", "Implementation", "1200", "Planned"],
    ["W11", "Nov 17", "Pinterest vs Instagram Reels: Which Has Better Download Options?", "Pinterest Tips", "Comparison (Both)", "pinterest vs instagram reels", "Consideration", "1800", "Planned"],
    ["W12", "Nov 24", "Pinterest Download FAQ: 15 Questions Answered", "Pinterest Downloading", "FAQ (Searchable)", "pinterest download faq", "Implementation", "2000", "Planned"],
]

for i, row in enumerate(calendar, 2):
    for j, val in enumerate(row, 1):
        ws3.cell(row=i, column=j, value=val)
    style_data(ws3, i, 9)
    # Alternate row colors by month
    week = row[0]
    if week in ["W1", "W2", "W3", "W4"]:
        for j in range(1, 10):
            ws3.cell(row=i, column=j).fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    elif week in ["W5", "W6", "W7", "W8"]:
        for j in range(1, 10):
            ws3.cell(row=i, column=j).fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    elif week in ["W9", "W10", "W11", "W12"]:
        for j in range(1, 10):
            ws3.cell(row=i, column=j).fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")

for col, width in enumerate([6, 12, 50, 22, 22, 40, 16, 12, 14], 1):
    ws3.column_dimensions[get_column_letter(col)].width = width

# ============================================================
# SHEET 4: Topic Clusters
# ============================================================
ws4 = wb.create_sheet("Topic Clusters")
ws4.sheet_properties.tabColor = "F59E0B"

headers4 = ["Hub (Pillar)", "Spoke (Article)", "Target Keyword", "Internal Links To", "Content Type"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, 5)

clusters = [
    # Cluster 1
    ["How to Download Pinterest Videos", "Complete Guide to Downloading Pinterest Videos", "how to download pinterest videos", "/pinterest-video-downloader", "Hub"],
    ["", "Download Pinterest Videos on iPhone", "download pinterest videos iphone", "Hub article, /pinterest-video-downloader", "Spoke"],
    ["", "Download Pinterest Videos on Android", "download pinterest videos android", "Hub article, /pinterest-video-downloader", "Spoke"],
    ["", "Download Pinterest Videos on PC/Mac", "download pinterest videos pc", "Hub article, /pinterest-video-downloader", "Spoke"],
    ["", "Save Pinterest Videos to Camera Roll", "save pinterest videos camera roll", "Hub article, /pinterest-video-downloader", "Spoke"],
    ["", "Pinterest Video Quality Guide", "pinterest video quality", "Hub article", "Spoke"],

    # Cluster 2
    ["Best Pinterest Downloaders", "SavePin vs Other Pinterest Downloaders", "best pinterest downloader", "/pinterest-video-downloader", "Hub"],
    ["", "Free Pinterest Downloaders Ranked", "free pinterest downloader", "Hub article", "Spoke"],
    ["", "Pinterest Downloader Apps Compared", "pinterest downloader app", "Hub article", "Spoke"],
    ["", "Pinterest Chrome Extensions Review", "pinterest downloader chrome", "Hub article", "Spoke"],

    # Cluster 3
    ["Pinterest Content Types", "How to Download Pinterest Images", "download pinterest images", "/pinterest-image-downloader", "Hub"],
    ["", "How to Download Pinterest GIFs", "download pinterest gifs", "/pinterest-gif-downloader", "Spoke"],
    ["", "How to Download Pinterest Reels", "download pinterest reels", "/pinterest-video-downloader", "Spoke"],
    ["", "Pinterest Board Downloader", "pinterest board download", "Hub article", "Spoke"],

    # Cluster 4
    ["Pinterest for Business", "Pinterest Marketing Guide 2026", "pinterest marketing 2026", "Hub article", "Hub"],
    ["", "Pinterest SEO: How to Get Found", "pinterest seo", "Hub article", "Spoke"],
    ["", "Pinterest vs Instagram for Business", "pinterest vs instagram", "Hub article", "Spoke"],

    # Cluster 5
    ["Legal & Ethics", "Is It Legal to Download Pinterest Videos?", "is it legal to download pinterest", "/terms", "Hub"],
    ["", "Pinterest Copyright Explained", "pinterest copyright", "Hub article, /privacy-policy", "Spoke"],
    ["", "Fair Use and Pinterest Content", "fair use pinterest", "Hub article", "Spoke"],
]

for i, row in enumerate(clusters, 2):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val)
    style_data(ws4, i, 5)
    if row[4] == "Hub":
        ws4.cell(row=i, column=1).fill = PILLAR_FILL
        ws4.cell(row=i, column=1).font = BOLD

for col, width in enumerate([30, 45, 35, 40, 12], 1):
    ws4.column_dimensions[get_column_letter(col)].width = width

# ============================================================
# SHEET 5: Content Scorecard
# ============================================================
ws5 = wb.create_sheet("Content Scorecard")
ws5.sheet_properties.tabColor = "8B5CF6"

headers5 = ["Content Idea", "Customer Impact (40%)", "Content-Market Fit (30%)", "Search Potential (20%)", "Resources (10%)", "Total Score", "Priority"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, 7)

ideas = [
    ["SavePin vs Pinterest Downloaders Comparison", 9, 10, 9, 8, None, "P0"],
    ["How to Download Pinterest Videos (Complete Guide)", 10, 10, 10, 7, None, "P0"],
    ["Free Pinterest Downloaders Ranked & Tested", 8, 9, 9, 7, None, "P0"],
    ["Download Pinterest Videos on iPhone", 8, 9, 8, 8, None, "P1"],
    ["Download Pinterest Videos on Android", 8, 9, 8, 8, None, "P1"],
    ["Is It Legal to Download Pinterest Videos?", 7, 8, 7, 6, None, "P1"],
    ["Pinterest Video Quality Explained", 6, 7, 6, 5, None, "P1"],
    ["Save Pinterest Videos to Camera Roll", 7, 8, 7, 7, None, "P1"],
    ["Pinterest Download Statistics 2026", 5, 6, 7, 5, None, "P1"],
    ["Pinterest Downloader Chrome Extensions", 6, 7, 6, 6, None, "P2"],
    ["Pinterest vs Instagram Reels Comparison", 6, 6, 6, 5, None, "P2"],
    ["Pinterest Marketing Guide 2026", 7, 6, 8, 6, None, "P2"],
    ["Pinterest Board Downloader Guide", 5, 7, 5, 5, None, "P2"],
    ["Pinterest Download FAQ", 6, 7, 5, 4, None, "P2"],
]

for i, row in enumerate(ideas, 2):
    # Calculate total score
    row[5] = round(row[1] * 0.4 + row[2] * 0.3 + row[3] * 0.2 + row[4] * 0.1, 1)
    for j, val in enumerate(row, 1):
        ws5.cell(row=i, column=j, value=val)
    style_data(ws5, i, 7)
    # Color by priority
    if row[6] == "P0":
        ws5.cell(row=i, column=7).fill = HIGH_FILL
    elif row[6] == "P1":
        ws5.cell(row=i, column=7).fill = MED_FILL
    else:
        ws5.cell(row=i, column=7).fill = LOW_FILL
    # Color score
    score = row[5]
    if score >= 8.5:
        ws5.cell(row=i, column=6).fill = HIGH_FILL
    elif score >= 7.0:
        ws5.cell(row=i, column=6).fill = MED_FILL

for col, width in enumerate([45, 18, 22, 18, 14, 14, 10], 1):
    ws5.column_dimensions[get_column_letter(col)].width = width

# ============================================================
# SHEET 6: Distribution Plan
# ============================================================
ws6 = wb.create_sheet("Distribution Plan")
ws6.sheet_properties.tabColor = "06B6D4"

headers6 = ["Channel", "Platform", "Frequency", "Content Type", "Goal", "ORB Stage"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, 6)

dist = [
    ["Blog", "npftas.xyz/blog", "4x/month", "All content types", "SEO traffic, authority", "Owned"],
    ["Twitter/X", "@SavePinTool", "3x/week", "Tips, stats, thread versions of posts", "Discovery, engagement", "Rented"],
    ["Reddit", "r/pinterest, r/ios, r/android", "2x/week", "Helpful answers, tutorials", "Discovery, trust", "Borrowed"],
    ["Google Search Console", "npftas.xyz", "Ongoing", "Monitor indexing, queries", "SEO optimization", "Owned"],
    ["Bing Webmaster Tools", "npftas.xyz", "Ongoing", "Submit sitemap, monitor", "Bing visibility", "Owned"],
    ["Pinterest", "SavePin account", "3x/week", "Infographics, tool previews", "Brand awareness", "Rented"],
    ["YouTube", "SavePin channel", "1x/month", "Screen recording tutorials", "Long-tail search", "Rented"],
    ["Quora", "SavePin answers", "2x/week", "Answer download questions", "Discovery, trust", "Borrowed"],
]

for i, row in enumerate(dist, 2):
    for j, val in enumerate(row, 1):
        ws6.cell(row=i, column=j, value=val)
    style_data(ws6, i, 6)
    stage = row[5]
    if stage == "Owned":
        ws6.cell(row=i, column=6).fill = HIGH_FILL
    elif stage == "Rented":
        ws6.cell(row=i, column=6).fill = MED_FILL
    else:
        ws6.cell(row=i, column=6).fill = LOW_FILL

for col, width in enumerate([18, 25, 14, 40, 30, 14], 1):
    ws6.column_dimensions[get_column_letter(col)].width = width

# Save
wb.save("content-strategy.xlsx")
print("Content strategy Excel created successfully!")
